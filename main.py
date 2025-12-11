import telebot
import random
from telebot import types
from openpyxl import load_workbook
from config import TOKEN
import shutil
import os

bot = telebot.TeleBot(TOKEN)

products_data = {
    "GEPO-7": "GEPO-7  № 40 (UDXK (ursodezoksixol kislotasi) 100 mg; pol-pola ekstrakti 100 mg; silimarin 50 mg; L-glutamin 50 mg; L-sistein 50 mg L-glisin 50 mg. )",
    "GLIAL-MG": "GLIAL-MG   № 20 ( magniy L-treonat 100 mg; GAMK (gamma-amino moy kislota) 100 mg; melatonin 5 mg; vitamin B6 1,5 mg.)",
    "FERR-26": "FERR-26 № 30  (Temir-2 fumarat 100 mg; glisin 100 mg; vitamin C 100 mg; vitamin B2 1,5 mg; vitamin B6 1,5 mg. )",
    "SINORIN": "SINORIN 20 ml",
    "SINORIN KIDS": "SINORIN KIDS 15 ml"
}

product_price = {
    "GEPO-7": 134000,
    "GLIAL-MG": 90000,
    "FERR-26": 85000,
    "SINORIN": 40000,
    "SINORIN KIDS": 40000
}

USER_SELECTED = {}  # chat_id: {product: amount}
USER_STATE = {}     # chat_id: hozir qaysi mahsulotga son kiritilmoqda


def menu_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    row = []
    for i, product in enumerate(products_data.keys(), start=1):
        row.append(types.KeyboardButton(product))
        if i % 2 == 0:
            markup.row(*row)
            row = []
    if row:
        markup.row(*row)
    markup.row(types.KeyboardButton("📄 Spesifikatsiyani olish"))
    return markup


@bot.message_handler(commands=['start'])
def start(msg):
    chat_id = msg.chat.id
    USER_SELECTED[chat_id] = {}
    USER_STATE.pop(chat_id, None)
    bot.send_message(
        chat_id,
        "📝 Spesifikatsiya tayyorlash uchun!\n\nMahsulotlardan tanlang:",
        reply_markup=menu_keyboard()
    )


@bot.message_handler(func=lambda m: m.text in products_data.keys())
def product_selected(msg):
    chat_id = msg.chat.id
    product = msg.text

    if len(USER_SELECTED.get(chat_id, {})) >= 4:
        return bot.send_message(chat_id, "❗ Maksimal 4 ta mahsulot tanlashingiz mumkin!")

    USER_STATE[chat_id] = product
    bot.send_message(chat_id, f"«{product}» — necha dona kerak?")


@bot.message_handler(func=lambda m: m.chat.id in USER_STATE)
def amount_entered(msg):
    chat_id = msg.chat.id
    product = USER_STATE[chat_id]

    try:
        amount = float(msg.text.replace(',', '.'))
    except ValueError:
        return bot.send_message(chat_id, "❗ Iltimos faqat son kiriting!")

    USER_SELECTED.setdefault(chat_id, {})
    USER_SELECTED[chat_id][product] = amount
    USER_STATE.pop(chat_id)

    bot.send_message(
        chat_id,
        f"✔ «{product}» — {amount} ta qo‘shildi.\nYana mahsulot tanlashingiz mumkin yoki «📄 Spesifikatsiyani olish» tugmasini bosing.",
        reply_markup=menu_keyboard()
    )


@bot.message_handler(func=lambda m: m.text == "📄 Spesifikatsiyani olish")
def send_excel(msg):
    chat_id = msg.chat.id
    selected = USER_SELECTED.get(chat_id, {})
    chat_prefix = str(random.randint(300, 9999))

    if not selected:
        return bot.send_message(chat_id, "❗ Avval hech bo‘lmasa bitta mahsulot tanlang.")

    original_file = os.path.join(os.getcwd(), "alsi.xlsx")
    filename = os.path.join(os.getcwd(), f"ALSI_SPES_{chat_prefix}.xlsx")

    if os.path.exists(filename):
        os.remove(filename)

    shutil.copy(original_file, filename)
    wb = load_workbook(filename)
    ws = wb.active

    start_row = 5
    used_rows = []

    # Tanlangan mahsulotlarni yozish
    for i, (product, amount) in enumerate(selected.items(), start=start_row):
        ws[f"C{i}"].value = products_data[product]
        ws[f"E{i}"].value = amount
        ws[f"F{i}"].value = product_price[product]
        used_rows.append(i)

    # Bo‘sh qatorlarni o‘chirish (faqat tanlanmaganlar)
    # Yuqoridan pastga qarab o‘chiramiz
    for row in reversed(range(start_row, start_row + 6)):
        if row not in used_rows:
            ws.delete_rows(row)

    # Summa qatori — tanlangan mahsulotlardan keyingi qator
    sum_row = start_row + len(selected)
    ws[f"E{sum_row}"].value = f"=SUM(E{start_row}:E{sum_row-1})"
    ws[f"F{sum_row}"].value = f"=SUM(F{start_row}:F{sum_row-1})"
    ws[f"G{sum_row}"].value = f"=SUM(G{start_row}:G{sum_row-1})"
    ws[f"H{sum_row}"].value = f"=SUM(H{start_row}:H{sum_row-1})"

    # E2 yacheykadagi mavjud matnni olish
    old_value = ws['E2'].value or ""

    # E2 yacheykasiga yangi matn yozish (avvalgi matnning davomida)
    ws['E2'].value = old_value + chat_prefix

    wb.save(filename)

    # Foydalanuvchiga yuborish
    with open(filename, "rb") as f:
        bot.send_document(chat_id, f, caption="📄 Tayyorlangan spesifikatsiya")

    os.remove(filename)
    USER_SELECTED[chat_id] = {}


bot.polling(none_stop=True)
